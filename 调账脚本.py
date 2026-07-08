import pandas as pd
import os
import sys

def generate_adjustment_template(material_file, output_file=None):
    """
    读取物料底表，提取调账模板所需列，生成调账模板。
    
    调账模板列：
    商戶提交時間, id, name,
    單杯裝環保紙袋, 雙杯裝環保紙袋, 紙袋,
    单杯無紡布袋（保溫）, 双杯無紡布袋（保溫）, 中号無紡布袋（保温）, 标准無紡布袋（保温）,
    紙漿杯托（飲品）, 膠袋, 小膠袋, 調賬金額
    
    物料底表 → 调账模板 列映射：
    下單日期 → 商戶提交時間
    門店ID → id
    門店名稱 → name
    單杯裝紙袋 → 單杯裝環保紙袋
    雙杯裝紙袋 → 雙杯裝環保紙袋
    紙袋 → 紙袋
    單杯無紡布袋 → 单杯無紡布袋（保溫）
    雙杯無紡布袋 → 双杯無紡布袋（保溫）
    3號無紡布袋 → 中号無紡布袋（保温）
    4號無紡布袋 → 标准無紡布袋（保温）
    紙漿杯2托 + 紙漿杯托-4托 → 紙漿杯托（飲品）
    大膠袋 → 膠袋
    小膠袋 → 小膠袋
    """
    
    # 读取物料底表
    df = pd.read_excel(material_file)
    
    # 确保输出目录存在
    if output_file is None:
        output_dir = os.path.join(os.path.expanduser("~"), "Desktop", "Catdesk file")
        os.makedirs(output_dir, exist_ok=True)
        output_file = os.path.join(output_dir, "物資調賬模版表_生成.xlsx")
    
    # 构建调账模板 DataFrame
    result = pd.DataFrame()
    
    # A/B/C 列直接提取
    result["商戶提交時間"] = df["下單日期"]
    result["id"] = df["門店ID"]
    result["name"] = df["門店名稱"]
    
    # 物料数量列映射（NaN 保持为 NaN，有值的直接复制）
    result["單杯裝環保紙袋"] = df["單杯裝紙袋"]
    result["雙杯裝環保紙袋"] = df["雙杯裝紙袋"]
    result["紙袋"] = df["紙袋"]
    result["单杯無紡布袋（保溫）"] = df["單杯無紡布袋"]
    result["双杯無紡布袋（保溫）"] = df["雙杯無紡布袋"]
    result["中号無紡布袋（保温）"] = df["3號無紡布袋"]
    result["标准無紡布袋（保温）"] = df["4號無紡布袋"]
    
    # 紙漿杯托（飲品）= 紙漿杯2托 + 紙漿杯托-4托（两者都可能存在，合并）
    cup2 = df["紙漿杯2托"].fillna(0)
    cup4 = df["紙漿杯托-4托"].fillna(0)
    combined_cup = cup2 + cup4
    result["紙漿杯托（飲品）"] = combined_cup.replace(0, pd.NA)
    
    result["膠袋"] = df["大膠袋"]
    result["小膠袋"] = df["小膠袋"]
    
    # 調賬金額列：先佔位，後面用 openpyxl 寫入公式
    result["調賬金額"] = ""
    
    # 写入 Excel（先寫入數據，再用 openpyxl 添加公式）
    result.to_excel(output_file, index=False, engine="openpyxl")
    
    # 使用 openpyxl 為「調賬金額」列添加公式
    from openpyxl import load_workbook
    wb = load_workbook(output_file)
    ws = wb.active
    
    # 找到「調賬金額」列的列號
    header_row = 1
    target_col = None
    for cell in ws[header_row]:
        if cell.value == "調賬金額":
            target_col = cell.column
            break
    
    if target_col:
        # D=單杯裝環保紙袋, E=雙杯裝環保紙袋, F=紙袋, G=单杯無紡布袋（保溫）, H=双杯無紡布袋（保溫）
        # I=中号無紡布袋（保温）, J=标准無紡布袋（保温）, K=紙漿杯托（飲品）, L=膠袋, M=小膠袋
        for row_idx in range(2, ws.max_row + 1):
            formula = f"=320*D{row_idx}+360*E{row_idx}+200*F{row_idx}+150*G{row_idx}+180*H{row_idx}+230*I{row_idx}+280*J{row_idx}+200*K{row_idx}+29*L{row_idx}+22*M{row_idx}"
            ws.cell(row=row_idx, column=target_col, value=formula)
    
    wb.save(output_file)
    print(f"调账模板已生成: {output_file}")
    print(f"共 {len(result)} 行数据")
    return output_file


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("=" * 50)
        print("调账脚本 - 用法说明")
        print("=" * 50)
        print()
        print("命令格式:")
        print("  python 调账脚本.py <物料底表.xlsx> [输出文件.xlsx]")
        print()
        print("参数说明:")
        print("  <物料底表.xlsx>  必填 - 物料底表文件路径")
        print("  [输出文件.xlsx]  可选 - 生成的调账模板保存路径")
        print("                   省略时默认保存到桌面 Catdesk file 文件夹")
        print()
        print("示例:")
        print('  python "调账脚本.py" "物料订单_20260615.xlsx"')
        print('  python "调账脚本.py" "物料订单_20260615.xlsx" "我的调账表.xlsx"')
        print()
        # 默认读取桌面路径
        default_input = os.path.join(os.path.expanduser("~"), "Desktop", "终版 单-物料底板", "物料订单_20260615.xlsx")
        if os.path.exists(default_input):
            print("检测到默认物料底表，正在生成...")
            generate_adjustment_template(default_input)
        else:
            print("未找到默认物料底表，请提供文件路径。")
    else:
        input_file = sys.argv[1]
        output_file = sys.argv[2] if len(sys.argv) > 2 else None
        generate_adjustment_template(input_file, output_file)
