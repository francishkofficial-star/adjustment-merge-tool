import pandas as pd
import os
from openpyxl import load_workbook

filepath = r'C:\Users\Francis LIU\Desktop\7.1-7.7調賬.xlsx'

# Read with openpyxl to see raw cell data including F-Q columns
wb = load_workbook(filepath, data_only=True)
sheet = wb.active
print(f"Sheet name: {sheet.title}")
print(f"Dimensions: {sheet.dimensions}")
print(f"Max row: {sheet.max_row}, Max col: {sheet.max_column}")

# Print first 5 rows all columns
for row_idx in range(1, min(6, sheet.max_row + 1)):
    row_data = []
    for col_idx in range(1, sheet.max_column + 1):
        cell = sheet.cell(row=row_idx, column=col_idx)
        col_letter = cell.column_letter
        row_data.append(f"{col_letter}={cell.value}")
    print(f"Row {row_idx}: {row_data}")

# Save full data to txt
with open(r'C:\Users\Francis LIU\Desktop\Catdesk file\preview_717.txt', 'w', encoding='utf-8') as f:
    f.write(f"Sheet: {sheet.title}\n")
    f.write(f"Rows: {sheet.max_row}, Cols: {sheet.max_column}\n\n")
    for row_idx in range(1, sheet.max_row + 1):
        row_data = []
        for col_idx in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            col_letter = cell.column_letter
            row_data.append(f"{col_letter}={cell.value}")
        f.write(f"Row {row_idx}: {row_data}\n")

# Also try with pandas, all columns
df = pd.read_excel(filepath)
df.to_csv(r'C:\Users\Francis LIU\Desktop\Catdesk file\preview_717.csv', index=False, encoding='utf-8-sig')
