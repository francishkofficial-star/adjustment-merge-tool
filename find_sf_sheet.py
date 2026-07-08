import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.load_workbook(r'C:\Users\Francis LIU\Desktop\Q1Q2预算总表2026-20260616（1）.xlsx')

# Look for a sheet with "顺丰" in the name or check all sheets for product codes
for sheet_name in wb.sheetnames:
    s = wb[sheet_name]
    # Search for "顺丰" or product codes like Keeta001 in headers or first few rows
    found = False
    for r in range(1, min(s.max_row + 1, 10)):
        for c in range(1, min(s.max_column + 1, 100)):
            v = s.cell(row=r, column=c).value
            if v and isinstance(v, str) and ('顺丰' in v or 'Keeta' in v):
                print(f"Sheet '{sheet_name}' row {r} col {c}: {v}")
                found = True
    if found:
        print(f"\n=== Full headers for sheet '{sheet_name}' ===")
        for c in range(1, min(s.max_column + 1, 100)):
            v = s.cell(row=1, column=c).value
            print(f"  Col {c}: {v}")
