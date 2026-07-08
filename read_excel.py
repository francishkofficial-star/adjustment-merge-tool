import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.load_workbook(r'C:\Users\Francis LIU\Desktop\Q1Q2预算总表2026-20260616（1）.xlsx')
print('Sheets:', wb.sheetnames)

for sheet_name in wb.sheetnames:
    s = wb[sheet_name]
    print(f"\n=== Sheet: {sheet_name} ===")
    print(f"Rows: {s.max_row}, Cols: {s.max_column}")
    # Print first row headers up to col 100
    headers = []
    for c in range(1, min(s.max_column + 1, 100)):
        v = s.cell(row=1, column=c).value
        headers.append((c, v))
    for c, v in headers:
        print(f"  Col {c}: {v}")
