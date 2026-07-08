import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
import os
from datetime import datetime

# 路径配置
high_end_path = r'C:\Users\Francis LIU\Desktop\高端0625.xlsx'
csp_path = r'C:\Users\Francis LIU\Desktop\CSP\CSP商品管理模板.xlsx'
vlookup_path = r'C:\Users\Francis LIU\Desktop\纯Vlookup.xlsx'
output_dir = r'C:\Users\Francis LIU\Desktop\Catdesk file'
output_path = os.path.join(output_dir, f'顺丰出库单{datetime.now().strftime("%Y-%m-%d")}.xlsx')

os.makedirs(output_dir, exist_ok=True)

# 读取数据
high_df = pd.read_excel(high_end_path)
csp_df = pd.read_excel(csp_path)
vlookup_df = pd.read_excel(vlookup_path)

# 构建商品映射：高端三合一(G列) -> 商品编码(B列)
mapping = {}
for _, row in csp_df.iterrows():
    code = str(row.iloc[1]).strip() if pd.notna(row.iloc[1]) else ''
    name_order = str(row.iloc[3]).strip() if pd.notna(row.iloc[3]) else ''
    name_high = str(row.iloc[6]).strip() if pd.notna(row.iloc[6]) else ''
    if name_high and code:
        mapping[name_high] = code
    if name_order and code:
        mapping[name_order] = code

# 构建Vlookup映射：门店id -> (contact, address, 门店名称)
vlookup_map = {}
for _, row in vlookup_df.iterrows():
    shop_id = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else ''
    contact = str(row.iloc[1]).strip() if pd.notna(row.iloc[1]) else ''
    address = str(row.iloc[2]).strip() if pd.notna(row.iloc[2]) else ''
    shop_name = str(row.iloc[3]).strip() if pd.notna(row.iloc[3]) else ''
    if shop_id:
        vlookup_map[shop_id] = (contact, address, shop_name)

# 手动补充未映射商品
mapping['普通紙袋（黄）'] = 'Keeta008'
mapping['高端纸袋（黑色普通）'] = 'Keeta011'
unmapped = set()

# 输出列定义（匹配顺丰出库单模板）
output_columns = [
    '仓库代码', '货主编码', '顺丰订单类型', '运费支付方式', '月结账号',
    'ERP行号', 'ERP订单号', '承运商', '承运商产品', '商品编码',
    '发货数量', '商品单位', '收方国家', '收件方省份', '收件方城市',
    '收件方区/县', '店铺名称', '收方公司', '收方姓名', '收方电话',
    '收方手机', '收件地址', '收件方电子邮箱', '收件方邮编', '收件方证件类型',
    '收件方证件号码', '是否需要保价', '保价金额', '是否需要代收货款', '代收货款金额',
    '代收货款月结卡号', '单价', '订单总金额', '币种简码', '海关编号',
    '商品品牌', '商品型号', '税金结算方式', '税金结算账号', '订单备注',
    '是否签回单', '是否自取件', '库存属性', '运单打印寄件方信息来源'
]

# 创建输出数据
output_rows = []
erp_counter = 1

for _, row in high_df.iterrows():
    shop_id = str(row['shopid']).strip() if pd.notna(row['shopid']) else ''
    material = str(row['提報物料']).strip() if pd.notna(row['提報物料']) else ''
    qty = int(row['物料數量']) if pd.notna(row['物料數量']) else 0
    
    product_code = mapping.get(material, '')
    if not product_code:
        unmapped.add(material)
        continue
    
    contact, address, shop_name = vlookup_map.get(shop_id, ('', '', ''))
    if not shop_name:
        shop_name = str(row['门店名称']).strip() if pd.notna(row['门店名称']) else ''
    
    output_row = {
        '仓库代码': '1006DCD',
        '货主编码': '8526947236',
        '顺丰订单类型': '销售订单',
        '运费支付方式': '寄付',
        '月结账号': '8526947236',
        'ERP行号': erp_counter,
        'ERP订单号': f'Keeta-{datetime.now().strftime("%Y-%m-%d")}-{erp_counter:04d}',
        '承运商': '顺丰速运',
        '承运商产品': '顺丰特快',
        '商品编码': product_code,
        '发货数量': qty,
        '商品单位': '包',
        '收方国家': '中国',
        '收件方省份': '香港特别行政区',
        '收件方城市': '香港',
        '收件方区/县': '',
        '店铺名称': shop_name,
        '收方公司': shop_id,
        '收方姓名': shop_name,
        '收方电话': contact,
        '收方手机': '',
        '收件地址': address,
        '收件方电子邮箱': '',
        '收件方邮编': '',
        '收件方证件类型': '',
        '收件方证件号码': '',
        '是否需要保价': '',
        '保价金额': '',
        '是否需要代收货款': '',
        '代收货款金额': '',
        '代收货款月结卡号': '',
        '单价': '',
        '订单总金额': '',
        '币种简码': '',
        '海关编号': '',
        '商品品牌': '',
        '商品型号': '',
        '税金结算方式': '',
        '税金结算账号': '',
        '订单备注': '',
        '是否签回单': '',
        '是否自取件': '',
        '库存属性': '',
        '运单打印寄件方信息来源': ''
    }
    output_rows.append(output_row)
    erp_counter += 1

# 创建Excel
wb = Workbook()
ws = wb.active
ws.title = '顺丰出库单'

# 写入表头
for col_idx, col_name in enumerate(output_columns, 1):
    cell = ws.cell(row=1, column=col_idx, value=col_name)
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center', vertical='center')

# 写入数据
for row_idx, row_data in enumerate(output_rows, 2):
    for col_idx, col_name in enumerate(output_columns, 1):
        ws.cell(row=row_idx, column=col_idx, value=row_data.get(col_name, ''))

# 设置列宽
for col_idx in range(1, len(output_columns) + 1):
    col_letter = get_column_letter(col_idx)
    ws.column_dimensions[col_letter].width = 15

# 冻结首行
ws.freeze_panes = 'A2'

# 保存
wb.save(output_path)

# 输出结果到文件避免编码问题
with open(os.path.join(output_dir, 'convert_result.txt'), 'w', encoding='utf-8') as f:
    f.write(f'输出行数: {len(output_rows)}\n')
    f.write(f'未映射商品: {unmapped}\n')
    f.write(f'已保存到: {output_path}\n')
