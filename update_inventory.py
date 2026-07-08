import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
import sys
sys.stdout.reconfigure(encoding='utf-8')

# 读取RealtimeInventory文件
inv_path = r'C:\Users\Francis LIU\Downloads\RealtimeInventory-1783414061359.xlsx'
df_inv = pd.read_excel(inv_path)

# 按商品编码分组，汇总可用库存数量
inventory_data = df_inv.groupby('商品编码')['可用库存数量'].sum().to_dict()

print("RealtimeInventory中的商品编码和可用库存:")
for code, qty in sorted(inventory_data.items()):
    print(f"  {code}: {qty}")

# 读取库存管理.xlsx
mgmt_path = r'C:\Users\Francis LIU\Desktop\库存管理.xlsx'
wb = openpyxl.load_workbook(mgmt_path)
ws = wb.active

# 建立商品编码到库存管理SKU的映射关系
# 根据RealtimeInventory中的商品名称和库存管理中的SKU进行匹配
code_to_sku = {
    'Keeta001': '塑料袋(大) 495*(295+65*2)mm',
    'Keeta002': '塑料袋(小) 400*(220+65*2)mm',
    'Keeta003': '规格一单杯装 120*105*300mm',
    'Keeta004': '规格二双杯装210*120*300mm',
    'Keeta005': '规格三中号袋255*160*280mm',
    'Keeta006': '规格四大号袋290*175*310mm',
    'Keeta007': '标准无纺布袋Preminum （保温）290*175*310mm',
    'Keeta008': '纸袋 310*290*175mm',
    'Keeta009': '单杯纸袋',
    'Keeta010': '双杯纸袋',
    'Keeta011': '高端物料紙袋（黑）- 每箱200個',
    'Keeta012': '麗新集團',
    'Keeta013': '机纸 17m长，宽55mm，外径40mm',
    'Keeta014': '包装纸箱 40*23*19cm',
    'Keeta015': '包装纸箱 30*23*19cm',
    'Keeta016': '包装纸箱 25*23*19cm',
    'Keeta018': '奉其奉纸浆杯托-4托 (300个/箱)200*5*200mm',
    'Keeta019': '封签 95*35mm',
    'Keeta020': '高端封签（黑） 95*35mm',
    'Keeta021': '高端物料封簽(聖誕版)（黑色)',
    'Keeta023': '海报25x35cm（英文版）',
    'Keeta024': '玻璃门贴',
    'Keeta025': '取餐区贴纸',
    'Keeta026': '按门贴',
    'Keeta027': '优惠贴（中文版）',
    'Keeta028': '优惠贴（英文版）',
    'Keeta030': '亚克力营业水牌',
    'Keeta033': '餐垫包装【纸盒】/箱出',
    'Keeta034': '餐具包【中式筷子勺】/箱出',
    'Keeta035': '餐具包【西式刀叉勺】/箱出',
    'Keeta036': '餐盒外壳【木质】/箱出',
    'Keeta037': '餐盒盖【木质】/箱出',
    'Keeta038': '餐盒内托【三格托】/箱出',
    'Keeta039': '新機',
    'Keeta040': '翻新POS機【商米】',
    'Keeta041': '翻新POS機【真唯】',
    'Keeta042': '充電線',
    'Keeta043': '充電頭',
}

# CG列的列号（第87列，索引86）
cg_col = 87

print(f"\nCG列号: {cg_col}")
print(f"CG列字母: {get_column_letter(cg_col)}")

# 更新CG列的可用库存
print("\n更新CG列可用库存:")
updated_count = 0

for row in range(1, ws.max_row + 1):
    sku_cell = ws.cell(row=row, column=4)  # D列 = SKU
    sku_value = sku_cell.value
    
    if sku_value and isinstance(sku_value, str):
        for code, mapped_sku in code_to_sku.items():
            if mapped_sku in sku_value or sku_value in mapped_sku:
                qty = inventory_data.get(code, 0)
                cg_cell = ws.cell(row=row, column=cg_col)
                cg_cell.value = qty
                print(f"  Row {row}: {code} -> {sku_value}, CG列更新为 {qty}")
                updated_count += 1
                break

print(f"\n共更新 {updated_count} 行")

# 保存文件到Catdesk file目录
import os
output_dir = r'C:\Users\Francis LIU\Desktop\Catdesk file'
os.makedirs(output_dir, exist_ok=True)
output_path = os.path.join(output_dir, '库存管理_已更新.xlsx')
wb.save(output_path)
print(f"\n文件已保存到: {output_path}")
