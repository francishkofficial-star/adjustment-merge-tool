import pandas as pd
import json
from openpyxl import load_workbook

filepath = r'C:\Users\Francis LIU\Desktop\7.1-7.7調賬.xlsx'
wb = load_workbook(filepath, data_only=True)
sheet = wb.active

# Column mapping: F-Q are material columns
MATERIAL_COLS = [
    ('F', 6, '大膠袋'),
    ('G', 7, '小膠袋'),
    ('H', 8, '單杯無紡布袋'),
    ('I', 9, '雙杯無紡布袋'),
    ('J', 10, '3號無紡布袋'),
    ('K', 11, '4號無紡布袋'),
    ('L', 12, '紙袋'),
    ('M', 13, '紙漿杯2托'),
    ('N', 14, '紙漿杯托-4托'),
    ('O', 15, '飲品封口紙'),
    ('P', 16, '雙杯裝紙袋'),
    ('Q', 17, '單杯裝紙袋'),
]

# Read all rows (skip header at row 1)
raw_rows = []
for row_idx in range(2, sheet.max_row + 1):
    shop_id = sheet.cell(row=row_idx, column=2).value  # B
    if not shop_id:
        continue
    shop_id = str(shop_id).strip()
    shop_name = sheet.cell(row=row_idx, column=3).value  # C
    phone = sheet.cell(row=row_idx, column=4).value  # D
    address = sheet.cell(row=row_idx, column=5).value  # E
    amount = sheet.cell(row=row_idx, column=18).value or 0  # R

    materials = {}
    for col_letter, col_idx, mat_name in MATERIAL_COLS:
        val = sheet.cell(row=row_idx, column=col_idx).value or 0
        materials[mat_name] = int(val) if val else 0

    raw_rows.append({
        'shop_id': shop_id,
        'shop_name': shop_name,
        'phone': phone,
        'address': address,
        'amount': float(amount),
        'materials': materials
    })

# Merge by shop_id
merged = {}
for row in raw_rows:
    sid = row['shop_id']
    if sid not in merged:
        merged[sid] = {
            'shop_id': sid,
            'shop_name': row['shop_name'],
            'phone': row['phone'],
            'address': row['address'],
            'amount': 0,
            'materials': {}
        }
    merged[sid]['amount'] += row['amount']
    for mat_name, qty in row['materials'].items():
        merged[sid]['materials'][mat_name] = merged[sid]['materials'].get(mat_name, 0) + qty

# Build output with remark
output = []
for sid in sorted(merged.keys(), key=lambda x: (len(x), x)):
    data = merged[sid]
    # Build remark: material*qty for materials with qty > 0
    remark_parts = []
    for col_letter, col_idx, mat_name in MATERIAL_COLS:
        qty = data['materials'].get(mat_name, 0)
        if qty > 0:
            remark_parts.append(f'{mat_name}*{qty}')
    remark = ', '.join(remark_parts)
    output.append({
        'shop_id': sid,
        'shop_name': data['shop_name'],
        'phone': data['phone'],
        'address': data['address'],
        'amount': data['amount'],
        'materials': data['materials'],
        'remark': remark
    })

# Save JSON for embedding
with open(r'C:\Users\Francis LIU\Desktop\Catdesk file\merged_717.json', 'w', encoding='utf-8') as f:
    json.dump(output, f, ensure_ascii=False, indent=2)

print(f'Raw rows: {len(raw_rows)}, Merged rows: {len(output)}')
# Find duplicates
from collections import Counter
id_counts = Counter(r['shop_id'] for r in raw_rows)
dups = {k: v for k, v in id_counts.items() if v > 1}
print(f'Duplicate IDs: {len(dups)}')
for k, v in dups.items():
    print(f'  {k}: {v} rows')
