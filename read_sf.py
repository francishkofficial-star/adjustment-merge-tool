from openpyxl import load_workbook
wb = load_workbook(r'C:\Users\Francis LIU\Downloads\已攬收 (1).xlsx')
ws = wb.active
print('Sheet:', ws.title)
print('Rows:', ws.max_row, 'Cols:', ws.max_column)
for r in range(1, min(4, ws.max_row+1)):
    row_data = [ws.cell(r, c).value for c in range(1, ws.max_column+1)]
    print(f'Row {r}:', row_data)
