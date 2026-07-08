import pandas as pd
import json
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

filepath = r'C:\Users\Francis LIU\Desktop\7.1-7.7調賬（未合并）.xlsx'
wb = load_workbook(filepath, data_only=True)
sheet = wb.active

MATERIAL_COLS = [
    ('F', 6, '大膠袋'),
    ('G', 7, '小膠袋'),
    ('H', 8, '單杯無紡布袋'),
    ('I', 9, '雙杯無紡布袋'),
    ('J', 10, '3號無紡布袋'),
    ('K', 11, '4號無紡布袋'),
    ('L', 12, '紙袋'),
    ('M', 13, '紙漿杯2托'),
    ('N', 14, '紙漿杯托-4托'),
    ('O', 15, '飲品封口紙'),
    ('P', 16, '雙杯裝紙袋'),
    ('Q', 17, '單杯裝紙袋'),
]

# Read all rows
raw_rows = []
for row_idx in range(2, sheet.max_row + 1):
    submit_date = sheet.cell(row=row_idx, column=1).value  # A
    shop_id = sheet.cell(row=row_idx, column=2).value  # B
    if not shop_id:
        continue
    shop_id = str(shop_id).strip()
    shop_name = sheet.cell(row=row_idx, column=3).value  # C
    phone = sheet.cell(row=row_idx, column=4).value  # D
    address = sheet.cell(row=row_idx, column=5).value  # E

    materials = {}
    for col_letter, col_idx, mat_name in MATERIAL_COLS:
        val = sheet.cell(row=row_idx, column=col_idx).value or 0
        materials[mat_name] = int(val) if val else 0

    raw_rows.append({
        'submit_date': submit_date,
        'shop_id': shop_id,
        'shop_name': shop_name,
        'phone': phone,
        'address': address,
        'materials': materials
    })

# Merge by shop_id
merged = {}
for row in raw_rows:
    sid = row['shop_id']
    if sid not in merged:
        merged[sid] = {
            'shop_id': sid,
            'submit_date': row['submit_date'],
            'shop_name': row['shop_name'],
            'phone': row['phone'],
            'address': row['address'],
            'materials': {}
        }
        for col_letter, col_idx, mat_name in MATERIAL_COLS:
            merged[sid]['materials'][mat_name] = 0
    for mat_name, qty in row['materials'].items():
        merged[sid]['materials'][mat_name] += qty

# Build output
output = []
for sid in sorted(merged.keys(), key=lambda x: (len(x), x)):
    data = merged[sid]
    remark_parts = []
    for col_letter, col_idx, mat_name in MATERIAL_COLS:
        qty = data['materials'][mat_name]
        if qty > 0:
            remark_parts.append(f'{mat_name}*{qty}')
    remark = ', '.join(remark_parts)
    output.append({
        'shop_id': sid,
        'submit_date': data['submit_date'].strftime('%Y-%m-%d') if hasattr(data['submit_date'], 'strftime') else str(data['submit_date']) if data['submit_date'] else '',
        'shop_name': data['shop_name'],
        'phone': data['phone'],
        'address': data['address'],
        'materials': data['materials'],
        'remark': remark
    })

# Save JSON
with open(r'C:\Users\Francis LIU\Desktop\Catdesk file\merged_717_v2.json', 'w', encoding='utf-8') as f:
    json.dump(output, f, ensure_ascii=False, indent=2)

print(f'Raw rows: {len(raw_rows)}, Merged rows: {len(output)}')

# Also build Excel with formulas
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

wb_out = Workbook()
ws = wb_out.active
ws.title = '合併後數據'

# Header
headers = ['提交日期', '商家ID', '商家名稱', '收貨電話', '收貨地址']
for col_letter, col_idx, mat_name in MATERIAL_COLS:
    headers.append(mat_name)
headers.append('調賬費用')
headers.append('備註')

for ci, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=ci, value=h)
    cell.font = Font(bold=True, color='FFFFFF')
    cell.fill = PatternFill('solid', fgColor='2C3E50')
    cell.alignment = Alignment(horizontal='center')

# Price mapping
PRICES = {
    '大膠袋': 29,
    '小膠袋': 22,
    '單杯無紡布袋': 150,
    '雙杯無紡布袋': 180,
    '3號無紡布袋': 230,
    '4號無紡布袋': 280,
    '紙袋': 200,
    '紙漿杯2托': 200,
    '紙漿杯托-4托': 200,
    '飲品封口紙': 10,
    '雙杯裝紙袋': 360,
    '單杯裝紙袋': 320,
}

for ri, row in enumerate(output, 2):
    excel_row = ri
    # Write date as datetime object to prevent Excel serial number
    date_val = row['submit_date']
    if date_val and isinstance(date_val, str) and len(date_val) == 10:
        from datetime import datetime
        try:
            date_obj = datetime.strptime(date_val, '%Y-%m-%d')
            ws.cell(row=excel_row, column=1, value=date_obj)
        except:
            ws.cell(row=excel_row, column=1, value=date_val)
    else:
        ws.cell(row=excel_row, column=1, value=date_val)
    ws.cell(row=excel_row, column=2, value=row['shop_id'])
    ws.cell(row=excel_row, column=3, value=row['shop_name'])
    ws.cell(row=excel_row, column=4, value=row['phone'])
    ws.cell(row=excel_row, column=5, value=row['address'])
    
    col_offset = 6  # F starts at column 6
    for idx, (col_letter, col_idx, mat_name) in enumerate(MATERIAL_COLS):
        ws.cell(row=excel_row, column=col_offset + idx, value=row['materials'][mat_name])
    
    # Formula for amount: =F2*29+G2*22+...
    formula_parts = []
    for idx, (col_letter, col_idx, mat_name) in enumerate(MATERIAL_COLS):
        excel_col = get_column_letter(col_offset + idx)
        price = PRICES[mat_name]
        formula_parts.append(f"{excel_col}{excel_row}*{price}")
    formula = "=" + "+".join(formula_parts)
    amount_col = col_offset + len(MATERIAL_COLS)
    ws.cell(row=excel_row, column=amount_col, value=formula)
    
    remark_col = amount_col + 1
    ws.cell(row=excel_row, column=remark_col, value=row['remark'])

# Column widths
ws.column_dimensions['A'].width = 14
ws.column_dimensions['B'].width = 14
ws.column_dimensions['C'].width = 35
ws.column_dimensions['D'].width = 18
ws.column_dimensions['E'].width = 45
for i in range(6, 6 + len(MATERIAL_COLS)):
    ws.column_dimensions[get_column_letter(i)].width = 12
ws.column_dimensions[get_column_letter(6 + len(MATERIAL_COLS))].width = 12
ws.column_dimensions[get_column_letter(6 + len(MATERIAL_COLS) + 1)].width = 30

# Freeze header
ws.freeze_panes = 'A2'

wb_out.save(r'C:\Users\Francis LIU\Desktop\Catdesk file\合併後-物料調賬底表7.1-7.7.xlsx')
print('Excel saved with formulas.')
