import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.load_workbook(r'C:\Users\Francis LIU\Desktop\Q1Q2预算总表2026-20260616（1）.xlsx')
s = wb['每週物料庫存']

# Look for "順豐" in row 2 and find the column indices
print("=== Row 2 headers ===")
for c in range(1, min(s.max_column + 1, 100)):
    v = s.cell(row=2, column=c).value
    if v:
        print(f"  Col {c}: {v}")

# Also check row 1 for any 顺丰-related headers
print("\n=== Row 1 headers ===")
for c in range(1, min(s.max_column + 1, 100)):
    v = s.cell(row=1, column=c).value
    if v:
        print(f"  Col {c}: {v}")

# Search for "CG" in any cell
print("\n=== Searching for 'CG' ===")
for r in range(1, min(s.max_row + 1, 10)):
    for c in range(1, min(s.max_column + 1, 100)):
        v = s.cell(row=r, column=c).value
        if v and isinstance(v, str) and 'CG' in v:
            print(f"  Row {r} Col {c}: {v}")
