import sys, io, os, json, traceback
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8')

import openpyxl
from http.server import HTTPServer, BaseHTTPRequestHandler
import tempfile, shutil

HTML_CONTENT = '''<!DOCTYPE html>
<html lang="zh-HK">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>調賬底表合併 + 備註生成器</title>
<style>
* { margin: 0; padding: 0; box-sizing: border-box; }
body { font-family: 'Segoe UI', 'Microsoft JhengHei', sans-serif; background: #f0f2f5; color: #333; min-height: 100vh; padding: 20px; }
.container { max-width: 1300px; margin: 0 auto; }
h1 { text-align: center; font-size: 24px; margin-bottom: 6px; color: #1a1a2e; }
.subtitle { text-align: center; font-size: 13px; color: #888; margin-bottom: 24px; }
.panel { background: #fff; border-radius: 12px; padding: 24px; margin-bottom: 20px; box-shadow: 0 2px 8px rgba(0,0,0,0.06); }
.panel-title { font-size: 16px; font-weight: 600; margin-bottom: 16px; color: #1a1a2e; border-left: 4px solid #4361ee; padding-left: 10px; }
.upload-area { border: 2px dashed #ccc; border-radius: 8px; padding: 32px; text-align: center; cursor: pointer; transition: all 0.3s; background: #fafafa; }
.upload-area:hover, .upload-area.dragover { border-color: #4361ee; background: #eef2ff; }
.upload-area .icon { font-size: 40px; margin-bottom: 8px; }
.upload-area .text { font-size: 14px; color: #666; }
.upload-area .filename { font-size: 14px; color: #4361ee; font-weight: 600; margin-top: 8px; }
.btn { padding: 10px 28px; border: none; border-radius: 6px; font-size: 14px; font-weight: 600; cursor: pointer; transition: all 0.3s; }
.btn-primary { background: #4361ee; color: #fff; }
.btn-primary:hover { background: #3a56d4; }
.btn-primary:disabled { background: #aaa; cursor: not-allowed; }
.btn-success { background: #2ecc71; color: #fff; }
.btn-success:hover { background: #27ae60; }
.btn-success:disabled { background: #aaa; cursor: not-allowed; }
.stats { display: flex; gap: 16px; margin-bottom: 16px; flex-wrap: wrap; }
.stat-item { background: #f8f9fa; padding: 12px 20px; border-radius: 8px; text-align: center; min-width: 110px; }
.stat-item .num { font-size: 22px; font-weight: 700; color: #4361ee; }
.stat-item .label { font-size: 12px; color: #888; margin-top: 4px; }
.result-wrapper { overflow: auto; max-height: 500px; border: 1px solid #e0e0e0; border-radius: 8px; }
.result-table { width: 100%; border-collapse: collapse; font-size: 12px; white-space: nowrap; }
.result-table th { background: #4361ee; color: #fff; padding: 8px 10px; text-align: left; position: sticky; top: 0; z-index: 10; font-weight: 600; }
.result-table td { padding: 6px 10px; border-bottom: 1px solid #eee; }
.result-table tr:nth-child(even) { background: #f9f9f9; }
.result-table .merged-row { background: #e8f5e9 !important; }
.result-table .remark-cell { color: #e65100; font-weight: 600; }
.info-box { background: #e8f4fd; border: 1px solid #b3d9f2; border-radius: 8px; padding: 14px 18px; font-size: 13px; color: #1a5276; line-height: 1.8; }
.info-box strong { color: #154360; }
.preview-section { display: none; }
.preview-section.show { display: block; }
.merged-detail { background: #fff3e0; border: 1px solid #ffcc80; border-radius: 6px; padding: 12px 16px; font-size: 13px; color: #e65100; margin-bottom: 12px; max-height: 200px; overflow-y: auto; display: none; }
.merged-detail.show { display: block; }
.merged-detail .detail-item { padding: 2px 0; }
.status-msg { padding: 10px 16px; border-radius: 6px; font-size: 13px; margin-top: 10px; display: none; }
.status-msg.error { background: #ffe0e0; border: 1px solid #ff9999; color: #c0392b; display: block; }
.status-msg.success { background: #e0f5e0; border: 1px solid #66cc66; color: #1a7a1a; display: block; }
.progress-bar { width: 100%; height: 6px; background: #e0e0e0; border-radius: 3px; margin-top: 10px; display: none; }
.progress-bar.show { display: block; }
.progress-fill { height: 100%; background: #4361ee; border-radius: 3px; transition: width 0.3s; width: 0%; }
</style>
</head>
<body>
<div class="container">
  <h1>調賬底表合併 + 備註生成器</h1>
  <p class="subtitle">合併相同商家ID的行（物料數量 + 調賬費用求和），自動生成備註列</p>
  <div class="panel">
    <div class="panel-title">功能說明</div>
    <div class="info-box">
      <strong>處理流程：</strong>
      1. 讀取調賬底表，以B列（商家ID）為唯一鍵，相同ID的多行合併為一行<br>
      2. F～Q列物料數量疊加求和，R列（調賬費用）也疊加求和<br>
      3. 合併後保留首次出現的商家名稱、電話、地址等資訊<br>
      4. S1寫入【備註】，S2起逐行填入「物料名稱*數量」（如：紙袋*2），多物料用逗號分隔<br>
      5. 所有數值轉為數字格式，下載Excel保留原始列結構
    </div>
  </div>
  <div class="panel">
    <div class="panel-title">上傳調賬底表</div>
    <div class="upload-area" id="uploadArea">
      <div class="icon">📊</div>
      <div class="text">點擊或拖拽上傳 Excel 檔案（.xlsx / .xls）</div>
      <div class="filename" id="fileName"></div>
    </div>
    <input type="file" id="fileInput" accept=".xlsx,.xls" style="display:none">
    <div id="uploadStatus" class="status-msg" style="display:none;"></div>
    <div class="progress-bar" id="progressBar"><div class="progress-fill" id="progressFill"></div></div>
    <div style="margin-top:16px; text-align:center;">
      <button class="btn btn-primary" id="btnProcess" disabled>開始處理</button>
    </div>
  </div>
  <div class="panel preview-section" id="resultSection">
    <div class="panel-title">處理結果</div>
    <div class="stats" id="stats"></div>
    <div class="merged-detail" id="mergedDetail"></div>
    <div style="margin-bottom:12px;">
      <button class="btn btn-success" id="btnDownload" disabled>下載 Excel</button>
    </div>
    <div class="result-wrapper">
      <table class="result-table" id="resultTable"></table>
    </div>
  </div>
</div>
<script>
var selectedFile = null;
var downloadUrl = null;

var uploadArea = document.getElementById('uploadArea');
var fileInput = document.getElementById('fileInput');
var fileName = document.getElementById('fileName');
var btnProcess = document.getElementById('btnProcess');

uploadArea.addEventListener('click', function() { fileInput.click(); });
uploadArea.addEventListener('dragover', function(e) { e.preventDefault(); uploadArea.classList.add('dragover'); });
uploadArea.addEventListener('dragleave', function() { uploadArea.classList.remove('dragover'); });
uploadArea.addEventListener('drop', function(e) {
  e.preventDefault();
  uploadArea.classList.remove('dragover');
  if (e.dataTransfer.files.length) handleFile(e.dataTransfer.files[0]);
});
fileInput.addEventListener('change', function(e) { if (e.target.files.length) handleFile(e.target.files[0]); });

function showStatus(msg, type) {
  var el = document.getElementById('uploadStatus');
  el.textContent = msg;
  el.className = 'status-msg ' + type;
}
function hideStatus() {
  document.getElementById('uploadStatus').style.display = 'none';
}

function handleFile(file) {
  hideStatus();
  fileName.textContent = file.name;
  selectedFile = file;
  btnProcess.disabled = false;
  showStatus('\\u2713 檔案已選擇：' + file.name + '（' + (file.size / 1024).toFixed(1) + ' KB），點擊「開始處理」。', 'success');
}

btnProcess.addEventListener('click', function() {
  if (!selectedFile) return;
  btnProcess.disabled = true;
  btnProcess.textContent = '處理中...';
  showStatus('正在上傳和處理檔案，請稍候...', 'success');
  document.getElementById('progressBar').classList.add('show');
  document.getElementById('progressFill').style.width = '30%';

  var formData = new FormData();
  formData.append('file', selectedFile);

  var xhr = new XMLHttpRequest();
  xhr.open('POST', '/process', true);
  xhr.onload = function() {
    document.getElementById('progressFill').style.width = '100%';
    if (xhr.status === 200) {
      try {
        var resp = JSON.parse(xhr.responseText);
        if (resp.success) {
          renderResult(resp);
          downloadUrl = '/download/' + resp.downloadId;
          showStatus('\\u2713 處理完成！可以下載 Excel 或查看預覽。', 'success');
        } else {
          showStatus('處理失敗：' + resp.error, 'error');
        }
      } catch(e) {
        showStatus('解析回應失敗：' + e.message, 'error');
      }
    } else {
      showStatus('伺服器錯誤：' + xhr.status, 'error');
    }
    btnProcess.disabled = false;
    btnProcess.textContent = '開始處理';
    setTimeout(function() { document.getElementById('progressBar').classList.remove('show'); }, 1000);
  };
  xhr.onerror = function() {
    showStatus('網路錯誤，請重試。', 'error');
    btnProcess.disabled = false;
    btnProcess.textContent = '開始處理';
    document.getElementById('progressBar').classList.remove('show');
  };
  xhr.send(formData);
});

function renderResult(resp) {
  document.getElementById('stats').innerHTML =
    '<div class="stat-item"><div class="num">' + resp.originalRows + '</div><div class="label">原始行數</div></div>' +
    '<div class="stat-item"><div class="num">' + resp.mergedRows + '</div><div class="label">合併後行數</div></div>' +
    '<div class="stat-item"><div class="num">' + resp.mergedCount + '</div><div class="label">合併了' + resp.mergedCount + '個ID</div></div>' +
    '<div class="stat-item"><div class="num">' + resp.remarkCount + '</div><div class="label">有備註行數</div></div>';

  var detailEl = document.getElementById('mergedDetail');
  if (resp.mergedDetails && resp.mergedDetails.length > 0) {
    var html = '<strong>合併詳情：</strong><br>';
    for (var i = 0; i < resp.mergedDetails.length; i++) {
      html += '<div class="detail-item">' + resp.mergedDetails[i] + '</div>';
    }
    detailEl.innerHTML = html;
    detailEl.classList.add('show');
  } else {
    detailEl.classList.remove('show');
  }

  var headers = resp.headers;
  var rows = resp.rows;
  var previewCols = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18];
  var previewNames = [];
  for (var i = 0; i < previewCols.length; i++) {
    previewNames.push(headers[previewCols[i]] || String.fromCharCode(65 + previewCols[i]));
  }
  var html = '<thead><tr>';
  for (var i = 0; i < previewNames.length; i++) html += '<th>' + previewNames[i] + '</th>';
  html += '</tr></thead><tbody>';
  var showRows = Math.min(rows.length, 100);
  for (var r = 0; r < showRows; r++) {
    var row = rows[r];
    var isMerged = row._merged;
    html += '<tr class="' + (isMerged ? 'merged-row' : '') + '">';
    for (var c = 0; c < previewCols.length; c++) {
      var val = row[previewCols[c]];
      if (previewCols[c] === 18) html += '<td class="remark-cell">' + (val || '') + '</td>';
      else html += '<td>' + (val !== null && val !== undefined ? val : '') + '</td>';
    }
    html += '</tr>';
  }
  if (rows.length > 100) {
    html += '<tr><td colspan="' + previewCols.length + '" style="text-align:center;color:#999;padding:12px">... 還有 ' + (rows.length - 100) + ' 行（下載Excel查看全部）</td></tr>';
  }
  html += '</tbody>';
  document.getElementById('resultTable').innerHTML = html;

  document.getElementById('resultSection').classList.add('show');
  document.getElementById('btnDownload').disabled = false;
}

document.getElementById('btnDownload').addEventListener('click', function() {
  if (!downloadUrl) return;
  var a = document.createElement('a');
  a.href = downloadUrl;
  a.download = '';
  document.body.appendChild(a);
  a.click();
  document.body.removeChild(a);
});
</script>
</body>
</html>'''

PORT = 8765
output_files = {}

class Handler(BaseHTTPRequestHandler):
    def log_message(self, format, *args):
        pass  # suppress logs

    def do_GET(self):
        if self.path == '/' or self.path == '/index.html':
            self.send_response(200)
            self.send_header('Content-Type', 'text/html; charset=utf-8')
            self.end_headers()
            self.wfile.write(HTML_CONTENT.encode('utf-8'))
        elif self.path.startswith('/download/'):
            file_id = self.path.replace('/download/', '')
            if file_id in output_files:
                filepath = output_files[file_id]
                self.send_response(200)
                self.send_header('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                self.send_header('Content-Disposition', 'attachment; filename="调账底表_合并并添加备注.xlsx"')
                self.end_headers()
                with open(filepath, 'rb') as f:
                    self.wfile.write(f.read())
            else:
                self.send_response(404)
                self.end_headers()
        else:
            self.send_response(404)
            self.end_headers()

    def do_POST(self):
        if self.path == '/process':
            content_type = self.headers.get('Content-Type', '')
            if 'multipart/form-data' not in content_type:
                self.send_json({'success': False, 'error': 'Expected multipart form data'})
                return

            # Parse multipart form data manually
            boundary = content_type.split('boundary=')[1].encode()
            content_length = int(self.headers['Content-Length'])
            body = self.rfile.read(content_length)

            # Extract file content
            parts = body.split(b'--' + boundary)
            file_data = None
            for part in parts:
                if b'filename=' in part:
                    # Find the start of file data (after empty line)
                    header_end = part.find(b'\r\n\r\n')
                    if header_end > 0:
                        file_data = part[header_end+4:]
                        # Remove trailing \r\n
                        if file_data.endswith(b'\r\n'):
                            file_data = file_data[:-2]
                    break

            if not file_data:
                self.send_json({'success': False, 'error': 'No file found in upload'})
                return

            # Save to temp file
            tmpdir = tempfile.mkdtemp()
            input_path = os.path.join(tmpdir, 'input.xlsx')
            with open(input_path, 'wb') as f:
                f.write(file_data)

            try:
                result = process_excel(input_path, tmpdir)
                self.send_json(result)
            except Exception as e:
                tb = traceback.format_exc()
                self.send_json({'success': False, 'error': str(e), 'traceback': tb})
            finally:
                shutil.rmtree(tmpdir, ignore_errors=True)
        else:
            self.send_response(404)
            self.end_headers()

    def send_json(self, data):
        self.send_response(200)
        self.send_header('Content-Type', 'application/json; charset=utf-8')
        self.end_headers()
        self.wfile.write(json.dumps(data, ensure_ascii=False).encode('utf-8'))


def to_num(val):
    if val is None or val == '':
        return 0
    if isinstance(val, (int, float)):
        return val
    try:
        n = float(val)
        return int(n) if n == int(n) else n
    except (ValueError, TypeError):
        return 0


def process_excel(input_path, tmpdir):
    wb = openpyxl.load_workbook(input_path, data_only=True)
    ws = wb.active

    # Read headers (row 1)
    headers = []
    for c in range(1, 20):  # A to S
        headers.append(ws.cell(1, c).value or '')

    MAT_START = 6   # F = col 6
    MAT_END = 17    # Q = col 17
    ID_COL = 2      # B = col 2
    AMOUNT_COL = 18 # R = col 18
    NAME_COL = 3    # C = col 3

    # Read all data rows
    rows = []
    for r in range(2, ws.max_row + 1):
        row_id = ws.cell(r, ID_COL).value
        if row_id is None or row_id == '':
            continue
        row_data = []
        for c in range(1, 19):  # A to R
            row_data.append(ws.cell(r, c).value)
        rows.append(row_data)

    # Group by store ID
    store_map = {}
    order = []
    merged_details = []

    for idx, row in enumerate(rows):
        sid = str(row[ID_COL - 1])
        if sid not in store_map:
            new_row = list(row)
            # Convert material cols and amount to numbers
            for c in range(MAT_START, MAT_END + 1):
                new_row[c - 1] = to_num(new_row[c - 1])
            new_row[AMOUNT_COL - 1] = to_num(new_row[AMOUNT_COL - 1])
            store_map[sid] = {'row': new_row, 'count': 1, 'rowIdx': idx + 2}
            order.append(sid)
        else:
            existing = store_map[sid]
            for c in range(MAT_START, MAT_END + 1):
                existing['row'][c - 1] += to_num(row[c - 1])
            existing['row'][AMOUNT_COL - 1] += to_num(row[AMOUNT_COL - 1])
            existing['count'] += 1

            mats = []
            for c in range(MAT_START, MAT_END + 1):
                v = to_num(row[c - 1])
                if v > 0:
                    mats.append(f'{headers[c-1]}+{v}')
            name = existing['row'][NAME_COL - 1] or '?'
            merged_details.append(f'ID={sid} ({name})：第{existing["rowIdx"]}行 + 第{idx+2}行 → {", ".join(mats)}, 費用+{to_num(row[AMOUNT_COL-1])}')

    # Build output
    output_headers = list(headers)
    output_headers.append('【备注】')

    output_rows = []
    remark_count = 0
    for sid in order:
        entry = store_map[sid]
        out_row = list(entry['row'])
        # Add remark
        parts = []
        for c in range(MAT_START, MAT_END + 1):
            mat_name = headers[c - 1]
            qty = out_row[c - 1]
            if qty and qty != 0:
                parts.append(f'{mat_name}*{qty}')
        remark = ', '.join(parts) if parts else ''
        if remark:
            remark_count += 1
        out_row.append(remark)
        out_row.append(entry['count'] > 1)  # _merged flag
        output_rows.append(out_row)

    # Create output Excel
    output_path = os.path.join(tempfile.gettempdir(), f'output_{os.getpid()}.xlsx')
    out_wb = openpyxl.Workbook()
    out_ws = out_wb.active
    out_ws.title = 'Sheet1'

    # Write headers
    for c, h in enumerate(output_headers):
        out_ws.cell(1, c + 1).value = h

    # Write data
    for r, row in enumerate(output_rows):
        for c in range(len(output_headers)):
            out_ws.cell(r + 2, c + 1).value = row[c]

    # Auto column width
    for c in range(len(output_headers)):
        max_len = len(str(output_headers[c] or ''))
        for r in range(min(len(output_rows), 50)):
            val = str(output_rows[r][c] if output_rows[r][c] is not None else '')
            if len(val) > max_len:
                max_len = len(val)
        from openpyxl.utils import get_column_letter
        out_ws.column_dimensions[get_column_letter(c + 1)].width = min(max(max_len * 1.5, 8), 45)

    out_wb.save(output_path)

    import uuid
    download_id = str(uuid.uuid4())
    output_files[download_id] = output_path

    # Prepare JSON response (limit rows for preview)
    preview_rows = []
    for row in output_rows:
        r = {}
        for c in range(len(output_headers)):
            val = row[c]
            if isinstance(val, float) and val == int(val):
                val = int(val)
            r[c] = val
        r['_merged'] = row[-1]
        preview_rows.append(r)

    return {
        'success': True,
        'originalRows': len(rows),
        'mergedRows': len(output_rows),
        'mergedCount': len(merged_details),
        'remarkCount': remark_count,
        'headers': output_headers,
        'rows': preview_rows,
        'mergedDetails': merged_details,
        'downloadId': download_id
    }


def main():
    server = HTTPServer(('127.0.0.1', PORT), Handler)
    print(f'伺服器已啟動: http://127.0.0.1:{PORT}')
    print(f'請在瀏覽器打開上述網址使用工具')
    server.serve_forever()

if __name__ == '__main__':
    main()
