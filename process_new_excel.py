import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import openpyxl
from openpyxl.utils import get_column_letter

def to_num(val):
    if val is None or val == '':
        return 0
    if isinstance(val, (int, float)):
        return val
    try:
        n = float(val)
        return int(n) if n == int(n) else n
    except (ValueError, TypeError):
        return 0

# Read the file
wb = openpyxl.load_workbook(r'C:\Users\Francis LIU\Desktop\调账底表6.24-6.30.xlsx', data_only=True)
ws = wb.active

# Constants
MAT_START = 6   # F = col 6
MAT_END = 17    # Q = col 17
ID_COL = 2      # B = col 2
AMOUNT_COL = 18 # R = col 18
NAME_COL = 3    # C = col 3

# Read headers
headers = []
for c in range(1, ws.max_column + 1):
    headers.append(ws.cell(1, c).value)

# Read all data rows
rows = []
for r in range(2, ws.max_row + 1):
    row_id = ws.cell(r, ID_COL).value
    if row_id is None or row_id == '':
        continue
    row_data = []
    for c in range(1, ws.max_column + 1):
        row_data.append(ws.cell(r, c).value)
    rows.append(row_data)

print(f'Read {len(rows)} data rows')

# Group by store ID
store_map = {}
order = []
merged_details = []

for idx, row in enumerate(rows):
    sid = str(row[ID_COL - 1])
    if sid not in store_map:
        new_row = list(row)
        # Convert material cols and amount to numbers
        for c in range(MAT_START, MAT_END + 1):
            new_row[c - 1] = to_num(new_row[c - 1])
        new_row[AMOUNT_COL - 1] = to_num(new_row[AMOUNT_COL - 1])
        store_map[sid] = {'row': new_row, 'count': 1, 'rowIdx': idx + 2}
        order.append(sid)
    else:
        existing = store_map[sid]
        for c in range(MAT_START, MAT_END + 1):
            existing['row'][c - 1] += to_num(row[c - 1])
        existing['row'][AMOUNT_COL - 1] += to_num(row[AMOUNT_COL - 1])
        existing['count'] += 1

        mats = []
        for c in range(MAT_START, MAT_END + 1):
            v = to_num(row[c - 1])
            if v > 0:
                mats.append(f'{headers[c-1]}+{v}')
        name = existing['row'][NAME_COL - 1] or '?'
        merged_details.append(f'ID={sid} ({name}): row {existing["rowIdx"]} + row {idx+2} -> {", ".join(mats)}, amount+{to_num(row[AMOUNT_COL-1])}')

print(f'Unique IDs: {len(order)}, Merged rows: {len(merged_details)}')

# Build output
output_headers = list(headers)
# Ensure S column has 【备注】 header
output_headers[18] = '【备注】' if len(output_headers) <= 18 else output_headers[18]
if len(output_headers) <= 18:
    output_headers.append('【备注】')

output_rows = []
remark_count = 0
for sid in order:
    entry = store_map[sid]
    out_row = list(entry['row'])
    # Add remark
    parts = []
    for c in range(MAT_START, MAT_END + 1):
        mat_name = headers[c - 1]
        qty = out_row[c - 1]
        if qty and qty != 0:
            parts.append(f'{mat_name}*{qty}')
    remark = ', '.join(parts) if parts else ''
    if remark:
        remark_count += 1
    
    # Ensure row has enough columns
    while len(out_row) < len(output_headers):
        out_row.append('')
    out_row[18] = remark  # S column (index 18)
    output_rows.append(out_row)

print(f'Output rows: {len(output_rows)}, With remarks: {remark_count}')

# Create output Excel
out_wb = openpyxl.Workbook()
out_ws = out_wb.active
out_ws.title = 'Sheet2'

# Write headers
for c, h in enumerate(output_headers):
    out_ws.cell(1, c + 1).value = h

# Write data
for r, row in enumerate(output_rows):
    for c in range(len(output_headers)):
        out_ws.cell(r + 2, c + 1).value = row[c]

# Auto column width
for c in range(len(output_headers)):
    max_len = len(str(output_headers[c] or ''))
    for r in range(min(len(output_rows), 50)):
        val = str(output_rows[r][c] if output_rows[r][c] is not None else '')
        if len(val) > max_len:
            max_len = len(val)
    out_ws.column_dimensions[get_column_letter(c + 1)].width = min(max(max_len * 1.5, 8), 45)

output_path = r'C:\Users\Francis LIU\Desktop\调账底表6.24-6.30_合并.xlsx'
out_wb.save(output_path)
print(f'Saved to: {output_path}')

# Show some merged details
if merged_details:
    print()
    print('Merged details (first 10):')
    for d in merged_details[:10]:
        print(f'  {d}')
