import sys, io, glob
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import pandas as pd
from openpyxl import load_workbook

template_files = glob.glob(r'C:\Users\Francis LIU\Desktop\vlookup*.xlsx')
print("Found:", template_files)

filepath = template_files[0]
wb = load_workbook(filepath, data_only=True)
print(f"\nSheets: {wb.sheetnames}")

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n=== Sheet: {sheet_name} ===")
    print(f"Rows: {ws.max_row}, Cols: {ws.max_column}")
    for row in ws.iter_rows(min_row=1, max_row=min(10, ws.max_row), values_only=False):
        vals = [(c.value, c.coordinate) for c in row]
        print(vals)

print("\n\n=== pandas read (first sheet) ===")
all_sheets = pd.read_excel(filepath, sheet_name=None, dtype=str, header=None)
for name, df in all_sheets.items():
    print(f"\nSheet: {name}")
    print(f"Shape: {df.shape}")
    print(df.head(10).to_string())
