import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.load_workbook(r'C:\Users\Francis LIU\Desktop\Q1Q2预算总表2026-20260616（1）.xlsx')

# Search all sheets more broadly for product codes and CG column
for sheet_name in wb.sheetnames:
    s = wb[sheet_name]
    print(f"\n=== Sheet: {sheet_name} (Rows: {s.max_row}, Cols: {s.max_column}) ===")
    # Search first 5 rows and up to 100 columns for any Keeta or product code patterns
    for r in range(1, min(s.max_row + 1, 5)):
        row_vals = []
        for c in range(1, min(s.max_column + 1, 100)):
            v = s.cell(row=r, column=c).value
            if v:
                row_vals.append(f"Col{c}={v}")
        if row_vals:
            print(f"  Row {r}: {' | '.join(row_vals[:20])}")
