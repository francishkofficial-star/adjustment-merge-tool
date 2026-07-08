import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
import os

file_path = r'C:\Users\Francis LIU\Desktop\Catdesk file\顺丰出库单2026-06-25.xlsx'
csp_path = r'C:\Users\Francis LIU\Desktop\CSP\CSP商品管理模板.xlsx'
wb = load_workbook(file_path)
ws = wb.active

# 读取CSP文件构建商品编码->商品型号映射
csp_df = pd.read_excel(csp_path)
model_map = {}
for _, row in csp_df.iterrows():
    code = str(row.iloc[1]).strip() if pd.notna(row.iloc[1]) else ''
    name = str(row.iloc[2]).strip() if pd.notna(row.iloc[2]) else ''
    if code and name:
        model_map[code] = name

# 找到列索引
header = [cell.value for cell in ws[1]]
col_erp_order = header.index('ERP订单号') + 1
col_erp_line = header.index('ERP行号') + 1
col_company = header.index('收方公司') + 1
col_product_code = header.index('商品编码') + 1
col_model = header.index('商品型号') + 1

# 按收方公司分组，分配ERP订单号
shop_seen = {}
erp_seq = 127

for row_idx in range(2, ws.max_row + 1):
    shop_id = str(ws.cell(row=row_idx, column=col_company).value).strip()
    if shop_id not in shop_seen:
        erp_no = f'Keeta-2026-0625-{erp_seq:04d}'
        shop_seen[shop_id] = erp_no
        erp_seq += 1
    ws.cell(row=row_idx, column=col_erp_order, value=shop_seen[shop_id])
    # ERP行号按同一收方公司内重新编号
    line_no = 1
    for r in range(2, row_idx + 1):
        if str(ws.cell(row=r, column=col_company).value).strip() == shop_id:
            if r == row_idx:
                ws.cell(row=row_idx, column=col_erp_line, value=line_no)
            line_no += 1
    
    # 补充商品型号(AK列)
    product_code = str(ws.cell(row=row_idx, column=col_product_code).value).strip() if ws.cell(row=row_idx, column=col_product_code).value else ''
    model_name = model_map.get(product_code, '')
    ws.cell(row=row_idx, column=col_model, value=model_name)

wb.save(file_path)

# 验证
wb2 = load_workbook(file_path)
ws2 = wb2.active
results = []
for row_idx in range(2, ws2.max_row + 1):
    shop = ws2.cell(row=row_idx, column=col_company).value
    erp = ws2.cell(row=row_idx, column=col_erp_order).value
    line = ws2.cell(row=row_idx, column=col_erp_line).value
    pcode = ws2.cell(row=row_idx, column=col_product_code).value
    model = ws2.cell(row=row_idx, column=col_model).value
    results.append(f'Row {row_idx}: shop={shop}, ERP={erp}, line={line}, code={pcode}, model={model}')

with open(r'C:\Users\Francis LIU\Desktop\Catdesk file\verify_erp.txt', 'w', encoding='utf-8') as f:
    f.write('\n'.join(results))
    f.write(f'\n\nUnique ERP count: {erp_seq - 127}\n')
    f.write(f'Unique shop count: {len(shop_seen)}\n')
