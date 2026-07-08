import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import openpyxl
from openpyxl.utils import get_column_letter

# Read the file
wb = openpyxl.load_workbook(r'C:\Users\Francis LIU\Desktop\调账底表6.24-6.30.xlsx', data_only=True)
ws = wb.active

print(f'Sheet: {ws.title}')
print(f'Rows: {ws.max_row}, Cols: {ws.max_column}')
print()

# Read headers
headers = []
for c in range(1, ws.max_column + 1):
    val = ws.cell(1, c).value
    headers.append(val)
    print(f'Col {c} ({get_column_letter(c)}): {val}')

print()
print(f'Total data rows: {ws.max_row - 1}')

# Check first 5 data rows
print()
print('First 5 data rows:')
for r in range(2, min(7, ws.max_row + 1)):
    vals = [ws.cell(r, c).value for c in range(1, min(20, ws.max_column + 1))]
    print(f'  Row {r}: {vals}')
